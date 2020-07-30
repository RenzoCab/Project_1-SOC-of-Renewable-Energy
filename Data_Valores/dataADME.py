#!/usr/bin/env python
# coding: utf-8

# In[1]:


import requests # https://curl.trillworks.com/#python
from bs4 import BeautifulSoup
import datetime
from datetime import timedelta
import xlwt
import pandas as pd # https://pythonspot.com/read-excel-with-pandas/
from openpyxl import load_workbook # https://openpyxl.readthedocs.io/en/stable/ wb=load_workbook('Termicas_ADME.xlsx')
# https://stackoverflow.com/questions/34767174/how-to-write-data-into-existing-xlsx-file-which-has-multiple-sheets
from openpyxl import Workbook # https://openpyxl.readthedocs.io/en/stable/ wb = Workbook()
import xlsxwriter # https://xlsxwriter.readthedocs.io wb = xlsxwriter.Workbook('Termicas_ADME.xlsx')
import openpyxl

# Author: Renzo Caballero
# KAUST: King Abdullah University of Science and Technology
# email: renzo.caballerorosas@kaust.edu.sa caballerorenzo@hotmail.com
# Website: None.
# August 2019; Last revision: 17/08/2019.


# In[2]:


def dataADME(date,station): # https://adme.com.uy/mmee/cvr.php

    headers = {
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.92 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Referer': 'https://adme.com.uy/mmee/cvr.php?central1=108&rbtn=horario&dtIni=20190601&dtFin=20190601&consultar=Consultar',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-US,en;q=0.9',
    }

    params = (
    ('central1', station),
    ('rbtn', 'horario'),
    ('dtIni', date),
    ('dtFin', date),
    ('consultar', 'Consultar'),
    )    

    #NB. Original query string below. It seems impossible to parse and
    #reproduce query strings 100% accurately so the one below is given
    #in case the reproduced version is not "correct".
    # response = requests.get('https://adme.com.uy/mmee/cvr.php?central1=108&rbtn=horario&dtIni=20190601&dtFin=20190601&consultar=Consultar', headers=headers)

    ok = 0;
    
    while ok == 0:
        try:
            response = requests.get('https://adme.com.uy/mmee/cvr.php', headers=headers, params=params)
            ok = 1;
            while response.status_code != 200:
                print(response.status_code)
                response = requests.get('https://adme.com.uy/mmee/cvr.php', headers=headers, params=params)        
        except:
            ok = 0;
    
    return response

def excelDate(date):
    temp = datetime.datetime(1899, 12, 30)
    delta = date - temp
    return float(delta.days) + (float(delta.seconds) / 86400)

def callData(date,station,fileName,sheetName):
    
    stringDate= date.strftime("%Y%m%d")
    
    response = dataADME(stringDate,station)
    response_soup = BeautifulSoup(response.content, 'html.parser')
    allTd = response_soup.find_all('td')
    List = list(allTd)
    newRow = [None] * 25
    newRow[0] = excelDate(date)
    for i in range(7,len(List)):
        newRow[i-6] = float(str(List[i].string))
#     fileName = 'Termicas_ADME.xlsx'
#     sheetName = 'Ciclo Combinado'

    wb = openpyxl.load_workbook(filename = fileName)
    ws = wb[sheetName]
    row = ws.max_row + 1
    for col, entry in enumerate(newRow, start=1):
        ws.cell(row=row, column=col, value=entry)
    wb.save(fileName)
    print(date.strftime("%Y%m%d"),sheetName)


# In[3]:


date = datetime.datetime(2018,1,1)
today = date.today()
twoDaysAgo = today - timedelta(days=2)
stringTDA = twoDaysAgo.strftime("%Y-%m-%d")
weWrote = 1;


while date.strftime("%d-%m-%Y") != twoDaysAgo.strftime("%d-%m-%Y"):
    
    if weWrote == 1:
    
        sheet1 = pd.read_excel('Termicas_ADME.xlsx', sheet_name='Ciclo Combinado', usecols=[0])
        # (*) There 'Ciclo Combinado' is the name of the sheet.
        sheet2 = pd.read_excel('Termicas_ADME.xlsx', sheet_name='CTR', usecols=[0])
        sheet3 = pd.read_excel('Termicas_ADME.xlsx', sheet_name='Motores Central Batlle', usecols=[0])
        sheet4 = pd.read_excel('Termicas_ADME.xlsx', sheet_name='PTA 7 y 8', usecols=[0])
        sheet5 = pd.read_excel('Termicas_ADME.xlsx', sheet_name='Punta del Tigre', usecols=[0])
        sheet6 = pd.read_excel('Valores_de_Agua_ADME.xlsx', sheet_name='Gabriel Terra (Bonete)', usecols=[0])
        sheet7 = pd.read_excel('Valores_de_Agua_ADME.xlsx', sheet_name='Rincon de Baygorria', usecols=[0])
        sheet8 = pd.read_excel('Valores_de_Agua_ADME.xlsx', sheet_name='Constitucion (Palmar)', usecols=[0])
        sheet9 = pd.read_excel('Valores_de_Agua_ADME.xlsx', sheet_name='Salto Grande Uruguay', usecols=[0])
        weWrote = 0
    
    stringDate = date.strftime("%Y-%m-%d")
    excelDateVar = excelDate(date)

    if not any(sheet1['Ciclo Combinado'][:] ==  excelDateVar): # pd.Timestamp(stringDate)):
    # (*) But here, 'Ciclo Combinado' is the name of the column.
        callData(date,'120','Termicas_ADME.xlsx','Ciclo Combinado')
        weWrote = 1
    
    if not any(sheet2['CTR'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'109','Termicas_ADME.xlsx','CTR')
        weWrote = 1
        
    if not any(sheet3['Motores Central Batlle'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'28','Termicas_ADME.xlsx','Motores Central Batlle')
        weWrote = 1
        
    if not any(sheet4['PTA 7 y 8'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'108','Termicas_ADME.xlsx','PTA 7 y 8')
        weWrote = 1
        
    if not any(sheet5['Punta del Tigre'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'51','Termicas_ADME.xlsx','Punta del Tigre')
        weWrote = 1
        
    if not any(sheet6['Gabriel Terra (Bonete)'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'102','Valores_de_Agua_ADME.xlsx','Gabriel Terra (Bonete)')
        weWrote = 1
            
    if not any(sheet7['Rincon de Baygorria'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'100','Valores_de_Agua_ADME.xlsx','Rincon de Baygorria')
        weWrote = 1
                
    if not any(sheet8['Constitucion (Palmar)'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'101','Valores_de_Agua_ADME.xlsx','Constitucion (Palmar)')
        weWrote = 1
                    
    if not any(sheet9['Salto Grande Uruguay'][:] == excelDateVar): # pd.Timestamp(stringDate)):
        callData(date,'105','Valores_de_Agua_ADME.xlsx','Salto Grande Uruguay')
        weWrote = 1
        
    date = date + timedelta(days=1)


# In[ ]:


station = ['94','106','97','120','109','9','91','77','28','92','108','51','101','102','100','105']

stationName = ['ALUR S.A.','APR A','Bioener','Ciclo Combinado','CTR','Fenirol','Galofer','Liderdat',              'Motores Central Batlle','Ponlar','PTA 7 y 8','Punta del Tigre','Constitucion (Palmar)',              'Gabriel Terra (Bonete)','Rincon de Baygorria','Salto Grande Uruguay']

# 'ALUR S.A.' --> '94'
# 'APR A' --> '106'
# 'Bioener' --> '97'
# 'Ciclo Combinado' --> '120'
# 'CTR' --> '109'
# 'Fenirol' --> '9'
# 'Galofer' --> '91'
# 'Liderdat' --> '77'
# 'Motores Central Batlle' --> '28'
# 'Ponlar' --> '92'
# 'PTA 7 y 8' --> '108'
# 'Punta del Tigre' --> '51'
# 'Constitucion (Palmar)' --> '101'
# 'Gabriel Terra (Bonete)' --> '102'
# 'Rincon de Baygorria' --> '100'
# 'Salto Grande Uruguay' --> '105'

